"""
excel_writer.py - 地域おこし協力隊 出納簿Excelへの書き込み処理

出納簿の列構成（テンプレート準拠）:
  A(1): №   B(2): "令和"   C(3): 年(令和数字)   D(4): "年"
  E(5): 月(数字)   F(6): "月"   G(7): 日(数字)   H(8): "日"
  I(9): 事業名   J(10): 勘定科目   K-O(11-15): 摘要(merged)
  P(16): 取引先   Q(17): 収入金額   R(18): 支出金額
  S(19): 差引残高(既存数式あり・上書きしない)
"""
import shutil
from io import BytesIO
from datetime import datetime
from copy import copy

DATA_START_ROW = 4   # データ開始行（テンプレートは行4から）
MAX_DATA_ROW = 77    # デフォルト値（動的検出で上書きされる）


def reiwa_year(year: int) -> int:
    return year - 2018


def detect_data_range(ws):
    """
    出納簿の実際のデータ範囲を動的に検出する。
    - A列にNo.（整数）が入っている行をデータスロットとみなす
    - 合計行（P列に数式 "=N..." など）をスキャン終端とする
    Returns: (data_end_row, totals_row_or_None)
    """
    data_end = DATA_START_ROW - 1
    totals_row = None

    for row_num in range(DATA_START_ROW, DATA_START_ROW + 500):
        a_val = ws.cell(row=row_num, column=1).value
        p_val = ws.cell(row=row_num, column=16).value

        # 合計行の検出: P列またはL列に "=N..." "=SUM..." などの数式
        if isinstance(p_val, str) and p_val.startswith('='):
            totals_row = row_num
            break

        # A列が整数 = データスロット行
        if isinstance(a_val, (int, float)) and a_val > 0:
            data_end = row_num
        elif a_val is None and data_end >= DATA_START_ROW:
            # A列がNoneかつデータ行が始まっていたら終端（空白エリア）
            # ただし合計行の手前まで続くので少し先まで見る
            pass

    return data_end, totals_row


def find_first_empty_row(ws):
    """
    出納簿シートで最初の空スロット行を返す。
    ファイルサイズに依存せず動的に検出。
    P列（取引先）・Q列（収入）・R列（支出）がすべて空かどうかで判断。
    """
    data_end, totals_row = detect_data_range(ws)
    scan_end = (totals_row - 1) if totals_row else (data_end + 200)

    for row_num in range(DATA_START_ROW, scan_end + 1):
        a_val = ws.cell(row=row_num, column=1).value
        p_val = ws.cell(row=row_num, column=16).value
        q_val = ws.cell(row=row_num, column=17).value
        r_val = ws.cell(row=row_num, column=18).value

        # A列に数字がある（データスロット）かつP/Q/R全部空 → 空きスロット
        if isinstance(a_val, (int, float)) and a_val > 0:
            if p_val is None and q_val is None and r_val is None:
                return row_num

    # スロットが全部埋まっていたら合計行の直前に新行追加
    return (totals_row - 1) if totals_row else (data_end + 1)


def count_filled_rows(ws):
    """実際にデータが入っている行数をカウント"""
    _, totals_row = detect_data_range(ws)
    scan_end = (totals_row - 1) if totals_row else (DATA_START_ROW + 500)

    count = 0
    for row_num in range(DATA_START_ROW, scan_end + 1):
        p_val = ws.cell(row=row_num, column=16).value
        q_val = ws.cell(row=row_num, column=17).value
        r_val = ws.cell(row=row_num, column=18).value
        if p_val is not None or q_val is not None or r_val is not None:
            count += 1
    return count


def is_duplicate(ws, date_str: str, vendor: str, amount: int) -> bool:
    """
    同じ取引先＋金額の組み合わせが既に存在するか確認（日付・取引先・金額の3つ）
    P列（取引先）と R列（支出）が両方空の行に達したらスキャン終了
    """
    for row_num in range(DATA_START_ROW, MAX_DATA_ROW + 1):
        p_val = ws.cell(row=row_num, column=16).value  # P列: 取引先
        r_val = ws.cell(row=row_num, column=18).value  # R列: 支出金額
        q_val = ws.cell(row=row_num, column=17).value  # Q列: 収入金額

        # 未使用行に達したら終了
        if p_val is None and r_val is None and q_val is None:
            break

        if str(p_val) == str(vendor) and r_val == amount:
            ry = ws.cell(row=row_num, column=3).value
            mn = ws.cell(row=row_num, column=5).value
            dy = ws.cell(row=row_num, column=7).value
            if ry and mn and dy:
                try:
                    row_date = datetime(ry + 2018, mn, dy).strftime("%Y-%m-%d")
                    if row_date == date_str:
                        return True
                except (ValueError, TypeError):
                    pass
    return False


def copy_row_format(ws, source_row: int, target_row: int):
    """
    source_row の書式（フォント・罫線・塗り・配置・数値書式）を
    target_row の各セルにコピーする。
    マージセルの子セルはスキップ。
    """
    # マージ済みセルの範囲を取得（子セルへの書き込みを防ぐ）
    merged_ranges = ws.merged_cells.ranges
    merged_cells = set()
    for mr in merged_ranges:
        for row in mr.rows:
            for cell_coord in row:
                merged_cells.add((cell_coord[0], cell_coord[1]))

    for col in range(1, 22):  # A〜U列
        src = ws.cell(row=source_row, column=col)
        tgt = ws.cell(row=target_row, column=col)

        # マージセルの子セル（左上以外）はスキップ
        if (target_row, col) in merged_cells:
            # 左上セルかどうか確認
            is_top_left = False
            for mr in merged_ranges:
                if mr.min_row == target_row and mr.min_col == col:
                    is_top_left = True
                    break
            if not is_top_left:
                continue

        try:
            if src.font:
                tgt.font = copy(src.font)
            if src.border:
                tgt.border = copy(src.border)
            if src.fill:
                tgt.fill = copy(src.fill)
            if src.alignment:
                tgt.alignment = copy(src.alignment)
            if src.number_format:
                tgt.number_format = src.number_format
        except Exception:
            pass

    # 行の高さもコピー
    src_dim = ws.row_dimensions.get(source_row)
    if src_dim and src_dim.height:
        ws.row_dimensions[target_row].height = src_dim.height


def write_single_row(ws, row_num: int, data: dict):
    """
    出納簿の指定行にデータを書き込む
    ※S列（差引残高）は既に数式が入っているので触らない
    ※K-O列はmergedなのでK列(11)のみ書き込む
    """
    date_str = data.get("date", "")
    vendor = data.get("vendor", "不明")
    amount = data.get("amount", 0)
    memo = data.get("memo", "") or vendor
    kamoku = data.get("kamoku", "消耗品")
    jigyo = data.get("jigyo", "ミッション活動")

    # 日付パース
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        ry = reiwa_year(dt.year)
        mn = dt.month
        dy = dt.day
    except (ValueError, TypeError):
        # 日付が不正な場合はデフォルト値
        ry, mn, dy = 7, 4, 1

    # ===== 書式コピー: 直前の既存行から書式を引き継ぐ =====
    ref_row = row_num - 1 if row_num > DATA_START_ROW else DATA_START_ROW
    # ref_rowがデータスロット行か確認（A列に数字があるか）
    if ref_row >= DATA_START_ROW:
        a_val = ws.cell(row=ref_row, column=1).value
        if isinstance(a_val, (int, float)) and a_val > 0:
            copy_row_format(ws, ref_row, row_num)

    # テンプレート範囲外（74行超）の場合はNo./令和/数式を手動で設定
    if row_num > MAX_DATA_ROW:
        # No. を前の行から連番
        prev_no = ws.cell(row=row_num - 1, column=1).value
        new_no = (prev_no + 1) if isinstance(prev_no, int) else row_num - DATA_START_ROW + 1
        ws.cell(row=row_num, column=1, value=new_no)   # A: No.
        ws.cell(row=row_num, column=2, value="令和")    # B
        ws.cell(row=row_num, column=4, value="年")      # D
        ws.cell(row=row_num, column=6, value="月")      # F
        ws.cell(row=row_num, column=8, value="日")      # H
        # 残高数式
        ws.cell(row=row_num, column=19,
                value=f"=S{row_num-1}+Q{row_num}-R{row_num}")

    # 書き込み（既存のA, B, D, F, H列は触らない）
    ws.cell(row=row_num, column=3, value=ry)       # C: 令和年
    ws.cell(row=row_num, column=5, value=mn)       # E: 月
    ws.cell(row=row_num, column=7, value=dy)       # G: 日
    ws.cell(row=row_num, column=9, value=jigyo)    # I: 事業名
    ws.cell(row=row_num, column=10, value=kamoku)  # J: 勘定科目

    # K列（摘要）: merged K-O なのでK(11)のみ書く
    try:
        ws.cell(row=row_num, column=11, value=memo)
    except AttributeError:
        pass  # MergedCellの場合はスキップ

    ws.cell(row=row_num, column=16, value=vendor)  # P: 取引先
    # Q列（収入）: 経費なのでNone（空のまま）
    ws.cell(row=row_num, column=18, value=amount)  # R: 支出金額
    # S列（差引残高）: テンプレートの数式をそのまま使う（74行以内）


def _count_existing_image_slots(ws) -> int:
    """
    領収書シートに既に貼られている画像スロット数を検出する。
    A列・G列の "No.X" ラベルを数えて返す。
    """
    count = 0
    LEFT_COL_NUM = 1
    RIGHT_COL_NUM = 7
    for row_num in range(1, ws.max_row + 2):
        for col_num in [LEFT_COL_NUM, RIGHT_COL_NUM]:
            val = ws.cell(row=row_num, column=col_num).value
            if isinstance(val, str) and val.strip().startswith("No."):
                count += 1
    return count


def add_receipt_images_to_sheet(ws, images: list):
    """
    領収書シートにレシート画像を貼り付ける。
    既存の画像スロットを検出し、続きから追記する（重複防止）。

    レイアウト:
      - 2列（左: A列、右: G列）
      - 各スロット: No.ラベル(1行) + 画像エリア(25行) + 余白(2行) = 28行
      - 開始行: 2行目から

    Args:
        ws: 「領収書　月分」シートのワークシートオブジェクト
        images: [(no, img_bytes), ...] のリスト
    """
    try:
        from openpyxl.drawing.image import Image as XLImage
    except ImportError:
        return

    ROWS_PER_IMAGE = 28   # ラベル1行 + 画像25行 + 余白2行
    IMG_WIDTH  = 260      # px
    IMG_HEIGHT = 340      # px
    START_ROW  = 2
    LEFT_COL   = "A"
    RIGHT_COL  = "G"
    LEFT_COL_NUM  = 1
    RIGHT_COL_NUM = 7

    # 既存スロット数を取得して続きから配置
    existing_count = _count_existing_image_slots(ws)

    for i, (no, img_bytes) in enumerate(images):
        if not img_bytes:
            continue

        slot_idx  = existing_count + i
        row_group = slot_idx // 2    # 何段目か（0始まり）
        col_side  = slot_idx % 2     # 0=左, 1=右

        label_row  = START_ROW + row_group * ROWS_PER_IMAGE
        image_row  = label_row + 1
        col_letter = LEFT_COL  if col_side == 0 else RIGHT_COL
        label_col  = LEFT_COL_NUM if col_side == 0 else RIGHT_COL_NUM

        # No.ラベルを書き込む
        ws.cell(row=label_row, column=label_col, value=f"No.{no}")

        # 画像をJPEGに変換（形式問わず安定化）
        try:
            from PIL import Image as PILImage
            import io as _io
            pil_img = PILImage.open(_io.BytesIO(img_bytes))
            if pil_img.mode in ('RGBA', 'P', 'LA'):
                pil_img = pil_img.convert('RGB')
            buf = _io.BytesIO()
            pil_img.save(buf, format="JPEG", quality=85)
            img_bytes = buf.getvalue()
        except Exception:
            pass  # 変換失敗時はそのまま使用

        # 画像を貼り付け
        try:
            img = XLImage(BytesIO(img_bytes))
            img.width  = IMG_WIDTH
            img.height = IMG_HEIGHT
            img.anchor = f"{col_letter}{image_row}"
            ws.add_image(img)
        except Exception:
            pass


def clear_data_rows(ws):
    """
    出納簿の全データ行のデータ列をクリアする（全書き直しモード用）。
    クリア対象: C(3) E(5) G(7) I(9) J(10) K(11) P(16) Q(17) R(18)
    保持する列: A(No.) B(令和) D(年) F(月) H(日) S(残高数式)
    """
    _, totals_row = detect_data_range(ws)
    scan_end = (totals_row - 1) if totals_row else (DATA_START_ROW + 500)

    clear_cols = [3, 5, 7, 9, 10, 11, 16, 17, 18]

    for row_num in range(DATA_START_ROW, scan_end + 1):
        a_val = ws.cell(row=row_num, column=1).value
        if isinstance(a_val, (int, float)) and a_val > 0:
            for col in clear_cols:
                ws.cell(row=row_num, column=col).value = None


def sort_records_by_date(records: list) -> list:
    """
    records を日付の古い順にソートして返す
    日付がないものは末尾に
    """
    def sort_key(r):
        d = r.get("date", "")
        return d if d else "9999-99-99"
    return sorted(records, key=sort_key)


def _get_or_create_receipt_sheet(wb, receipt_sheet_option: str, new_sheet_name: str):
    """
    領収書シートを取得または新規作成して返す。

    receipt_sheet_option:
      "new"       → 新規シート作成（new_sheet_nameを使用）
      "<シート名>" → 既存のそのシートを使用

    Returns: (worksheet, sheet_name)
    """
    if receipt_sheet_option == "new":
        # 新規シート名を決定（重複時は連番）
        base_name = new_sheet_name.strip() if new_sheet_name.strip() else "領収書"
        name = base_name
        counter = 2
        while name in wb.sheetnames:
            name = f"{base_name}({counter})"
            counter += 1
        ws = wb.create_sheet(title=name)
        return ws, name
    else:
        # 指定された既存シートを使用
        if receipt_sheet_option in wb.sheetnames:
            return wb[receipt_sheet_option], receipt_sheet_option
        # 見つからない場合は新規作成
        name = receipt_sheet_option
        ws = wb.create_sheet(title=name)
        return ws, name


def write_receipts_to_excel(
    excel_bytes: bytes,
    records: list,
    images: list,
    receipt_sheet_option: str = "auto",
    new_sheet_name: str = "",
    skip_sort: bool = False,
    rewrite_all: bool = False,
) -> tuple:
    """
    複数の経費データを出納簿Excelに書き込み、
    更新済みExcelのbytesと処理結果を返す

    Args:
        excel_bytes:          既存の出納簿Excelファイルのbytes
        records:              書き込む経費データのリスト
                              rewrite_all=True の場合、_type="existing"|"new" フィールドを含む
        images:               [(no, img_bytes), ...] 領収書画像リスト（新規アイテム分のみ）
        receipt_sheet_option: "new"=新規タブ作成 / "auto"=既存タブ自動検出 / シート名=そのタブに続き書き
        new_sheet_name:       receipt_sheet_option="new" の時の新しいタブ名
        skip_sort:            True=日付ソートをスキップ（ユーザー指定順を維持）
        rewrite_all:          True=既存データをクリアしてから全件書き直し（重複チェックなし）

    Returns:
        (updated_excel_bytes, results_list)
        results_list: [{"no", "vendor", "amount", "status", "row"}, ...]
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(excel_bytes))

    if "出納簿" not in wb.sheetnames:
        raise ValueError("「出納簿」シートが見つかりません。正しい出納簿ファイルをアップロードしてください。")

    ws_d = wb["出納簿"]
    results = []
    added_images = []

    if rewrite_all:
        # ===== 全書き直しモード =====
        # 1. 全データ行をクリア
        clear_data_rows(ws_d)

        # 2. 新規アイテムの画像イテレータ
        img_iter = iter(images)

        # 3. 全件を指定順で書き込む（重複チェックなし）
        for data in records:
            vendor = data.get("vendor", "不明")
            amount = data.get("amount", 0)

            target_row = find_first_empty_row(ws_d)
            write_single_row(ws_d, target_row, data)
            no = ws_d.cell(row=target_row, column=1).value

            results.append({
                "no": no,
                "vendor": vendor,
                "amount": amount,
                "status": "追加",
                "row": target_row,
            })

            # 新規アイテムのみ画像を対応付け（_type="new"）
            if data.get("_type") == "new":
                try:
                    _, img_bytes = next(img_iter)
                    if img_bytes:
                        added_images.append((no, img_bytes))
                except StopIteration:
                    pass

    else:
        # ===== 通常モード（新規追加・重複チェックあり）=====
        sorted_records = records if skip_sort else sort_records_by_date(records)

        for data in sorted_records:
            vendor = data.get("vendor", "不明")
            amount = data.get("amount", 0)
            date_str = data.get("date", "")

            # 重複チェック
            if is_duplicate(ws_d, date_str, vendor, amount):
                results.append({
                    "no": None,
                    "vendor": vendor,
                    "amount": amount,
                    "status": "重複スキップ",
                    "row": None,
                })
                continue

            # 書き込み先の空行を探す（74行超えた場合も自動追記）
            target_row = find_first_empty_row(ws_d)

            # 書き込み実行
            write_single_row(ws_d, target_row, data)

            # この行の№を取得（A列）
            no = ws_d.cell(row=target_row, column=1).value

            results.append({
                "no": no,
                "vendor": vendor,
                "amount": amount,
                "status": "追加",
                "row": target_row,
            })

        # 通常モードの画像対応: results と images は同順・同数
        added_images = [
            (result["no"], img_bytes)
            for result, (_, img_bytes) in zip(results, images)
            if result["status"] == "追加" and result["no"] is not None
        ]

    # ===== 領収書シートへの画像貼り付け =====
    if added_images:
        if receipt_sheet_option == "auto":
            # 後方互換: 最初に見つかった「領収書」シートを使用
            receipt_sheet_option = next(
                (n for n in wb.sheetnames if "領収書" in n), "new"
            )

        ws_r, _ = _get_or_create_receipt_sheet(wb, receipt_sheet_option, new_sheet_name)
        add_receipt_images_to_sheet(ws_r, added_images)

    # BytesIOに保存して返す
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read(), results
